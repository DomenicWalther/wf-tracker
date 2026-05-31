import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\Domenic\Downloads\WARFRAME COMPLETION TRACKER - PUBLIC.xlsx',
    data_only=True
)

def str_val(v):
    if v is None or isinstance(v, bool):
        return None
    s = str(v).strip()
    return s if s else None

data = {}

# === QUESTS ===
ws = wb['QUESTS']
quests = []
for row in ws.iter_rows(min_row=5, values_only=True):
    n = str_val(row[0])
    if n and n != 'Quest Name':
        quests.append(n)
data['quests'] = quests

# === GEAR ===
ws = wb['GEAR']
gear = {'warframes': [], 'primaries': [], 'secondaries': [], 'melees': [],
        'companions': [], 'archwings': [], 'archwingWeapons': [], 'extras': []}
section_map = {
    'Warframe': 'warframes', 'Primary': 'primaries', 'Secondary': 'secondaries',
    'Melee': 'melees', 'Companions': 'companions', 'Archwing': 'archwings',
    'Archwing Weapons': 'archwingWeapons', 'Extra': 'extras'
}
current_section = None
for row in ws.iter_rows(min_row=5, values_only=True):
    name = str_val(row[0])
    if not name:
        continue
    if name in section_map:
        current_section = section_map[name]
        continue
    if current_section:
        is_founder = (len(row) > 12 and str_val(row[12]) == 'founder')
        gear[current_section].append({'name': name, 'isFounderOnly': is_founder})
data['gear'] = gear

# === LICH GEAR ===
ws = wb['LICH GEAR']
lich = {'kuva': [], 'tenet': [], 'coda': []}
current_type = None
type_map = {'Kuva': 'kuva', 'Tenet': 'tenet', 'Coda': 'coda'}
for row in ws.iter_rows(min_row=6, values_only=True):
    t = str_val(row[0])
    if t in type_map:
        current_type = type_map[t]
    weapon = str_val(row[2])
    if weapon and current_type:
        lich[current_type].append(weapon)
data['lichGear'] = lich

# === INCARNON ===
ws = wb['INCARNON']
incarnon_families = []
current_family = None
current_week = None
for row in ws.iter_rows(min_row=6, values_only=True):
    week_val = row[0]
    if week_val is not None and not isinstance(week_val, bool):
        try:
            current_week = int(week_val)
        except (ValueError, TypeError):
            pass
    family_name = str_val(row[2])
    if family_name:
        current_family = {'name': family_name, 'weapons': []}
        if current_week is not None:
            current_family['week'] = current_week
        incarnon_families.append(current_family)
    weapon = str_val(row[4])
    if weapon and current_family:
        current_family['weapons'].append(weapon)
data['incarnon'] = incarnon_families

# === ARCANES ===
ws = wb['ARCANES']
arcanes = {'warframe': [], 'weapon': [], 'amp': [], 'companion': [], 'other': []}
current_section = 'other'
section_map_arc = {
    'WARFRAME': 'warframe', 'WEAPON': 'weapon', 'AMP': 'amp',
    'COMPANION': 'companion', 'OPERATOR': 'other', 'DRIFTER': 'other',
    'OTHER': 'other', 'KITGUN': 'weapon', 'ZAW': 'weapon'
}
for row in ws.iter_rows(min_row=5, values_only=True):
    name = str_val(row[0])
    if name and name in section_map_arc:
        current_section = section_map_arc[name]
        continue
    arcane_name = str_val(row[1])
    if arcane_name and arcane_name != 'Arcane' and arcane_name != 'Progress':
        arcanes[current_section].append(arcane_name)
data['arcanes'] = arcanes

# === MODS ===
ws = wb['MODS']
mods = []
current_category = 'General'
for row in ws.iter_rows(min_row=7, values_only=True):
    name = str_val(row[0])
    if not name:
        continue
    # Category headers have no owned/rank data
    if name and row[1] is None and row[2] is None and row[3] is None:
        current_category = name
        continue
    max_rank = row[3]
    if isinstance(max_rank, (int, float)):
        mods.append({'name': name, 'maxRank': int(max_rank), 'category': current_category})
data['mods'] = mods

# === SUBSUME ===
ws = wb['SUBSUME']
subsume = []
for row in ws.iter_rows(min_row=7, values_only=True):
    wf = str_val(row[0])
    ability = str_val(row[1])
    if wf and ability:
        subsume.append({'warframe': wf, 'ability': ability})
data['subsume'] = subsume

# === RAILJACK ===
ws = wb['RAILJACK']
railjack = {'intrinsics': ['Tactical', 'Piloting', 'Gunnery', 'Engineering', 'Command'],
            'components': []}
current_house = None
for row in ws.iter_rows(min_row=8, values_only=True):
    house = str_val(row[0])
    if house:
        current_house = house
    component = str_val(row[1])
    if component and current_house:
        railjack['components'].append({
            'house': current_house,
            'component': component,
            'bonus': str_val(row[2]) or ''
        })
data['railjack'] = railjack

# === RELICS ===
ws = wb['RELICS']
relics = {'Lith': [], 'Meso': [], 'Neo': [], 'Axi': []}
current_era = None
for row in ws.iter_rows(min_row=6, values_only=True):
    era = str_val(row[0])
    if era in relics:
        current_era = era
    relic = str_val(row[1])
    if relic and current_era:
        relics[current_era].append(relic)
data['relics'] = relics

# === BLUEPRINTS ===
ws = wb['BLUEPRINTS']
blueprints = {}
current_category = None
current_vendor = None
for row in ws.iter_rows(min_row=5, values_only=True):
    col0 = str_val(row[0])
    col1 = str_val(row[1])
    col3 = row[3] if len(row) > 3 else None  # isOld flag (0 = current, None = old)

    if col0 and not col1:
        # Category header
        current_category = col0
        current_vendor = None
        if current_category not in blueprints:
            blueprints[current_category] = {}
        continue

    if not current_category:
        continue

    if col0 and col1:
        # Vendor + item
        current_vendor = col0
        if current_vendor not in blueprints[current_category]:
            blueprints[current_category][current_vendor] = []
        blueprints[current_category][current_vendor].append({
            'name': col1,
            'isOld': col3 is None
        })
    elif col1 and current_vendor:
        # Just item under current vendor
        blueprints[current_category][current_vendor].append({
            'name': col1,
            'isOld': col3 is None
        })
data['blueprints'] = blueprints

# === ITEMS ===
ws = wb['ITEMS']
items = {}
current_section = None
for row in ws.iter_rows(min_row=8, values_only=True):
    col0 = str_val(row[0])
    col2 = str_val(row[2]) if len(row) > 2 else None

    # Section header: text in col0, no checkbox in col1
    if col0 and not isinstance(row[1], bool):
        current_section = col0
        if current_section not in items:
            items[current_section] = []
        # Also check if there's an item alongside section header
        if col2:
            items[current_section].append(col0)
        continue

    if current_section:
        if col0:
            items[current_section].append(col0)
        if col2:
            items[current_section].append(col2)
data['items'] = items

# === COSMETICS ===
ws = wb['COSMETICS']
# Structure: main category at col0, sub-type at col0 sometimes, items in cols 1,3,5,7,9
cosmetics = {}
current_main = None
current_sub = None
for row in ws.iter_rows(min_row=5, values_only=True):
    col0 = str_val(row[0])

    if col0:
        # Check if this is a main category (all caps, no items in row 1-10)
        has_items = any(str_val(row[i]) for i in range(1, 11) if len(row) > i)
        if not has_items:
            current_main = col0
            current_sub = None
            if current_main not in cosmetics:
                cosmetics[current_main] = {}
            continue
        else:
            # It's a sub-type with items in same row
            current_sub = col0
            if current_main not in cosmetics:
                cosmetics[current_main] = {}
            if current_sub not in cosmetics[current_main]:
                cosmetics[current_main][current_sub] = []

    if current_main:
        sub = current_sub or 'General'
        if sub not in cosmetics[current_main]:
            cosmetics[current_main][sub] = []
        for i in range(1, 11, 2):
            if len(row) > i:
                item = str_val(row[i])
                if item:
                    cosmetics[current_main][sub].append(item)

data['cosmetics'] = cosmetics

# === COLLECTABLE ===
ws = wb['COLLECTABLE']
collectable = {}
header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
col_to_section = {}
for i, v in enumerate(header_row):
    sv = str_val(v)
    if sv and not sv.startswith('Data'):
        col_to_section[i] = sv
        collectable[sv] = []

for row in ws.iter_rows(min_row=6, values_only=True):
    for col_idx, sec_name in col_to_section.items():
        if col_idx < len(row):
            item = str_val(row[col_idx])
            if item:
                collectable[sec_name].append(item)
data['collectable'] = collectable

# === DECORATIONS ===
ws = wb['DECORATIONS']
decorations = {}
current_section = None
for row in ws.iter_rows(min_row=5, values_only=True):
    col0 = str_val(row[0])
    col1 = str_val(row[1]) if len(row) > 1 else None
    col3 = str_val(row[3]) if len(row) > 3 else None
    if col0:
        current_section = col0
        if current_section not in decorations:
            decorations[current_section] = []
    if current_section:
        if col1:
            decorations[current_section].append(col1)
        if col3:
            decorations[current_section].append(col3)
data['decorations'] = decorations

# === CODEX ===
ws = wb['CODEX']
codex = {}
header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
col_to_codex = {}
for i, v in enumerate(header_row):
    sv = str_val(v)
    if sv:
        col_to_codex[i] = sv
        codex[sv] = []

for row in ws.iter_rows(min_row=7, values_only=True):
    for col_idx, sec_name in col_to_codex.items():
        if col_idx < len(row):
            item = str_val(row[col_idx])
            if item and item != 'ITEM NAME':
                codex[sec_name].append(item)
data['codex'] = codex

# === MARKET ===
ws = wb['MARKET']
market = {}
current_cat = None
for row in ws.iter_rows(min_row=5, values_only=True):
    col0 = str_val(row[0])
    if col0:
        current_cat = col0
        if current_cat not in market:
            market[current_cat] = []
    if not current_cat:
        continue
    for i in range(1, 7, 2):
        item = str_val(row[i]) if len(row) > i else None
        if item:
            market[current_cat].append(item)
data['market'] = market

# === EXTRA ===
ws = wb['EXTRA']
extra = {}
current_section = None
for row in ws.iter_rows(min_row=5, values_only=True):
    col0 = str_val(row[0])

    if col0 and '--->' in col0:
        current_section = col0.replace(' --->', '').replace('--->', '').strip()
        if current_section not in extra:
            extra[current_section] = []

    if current_section:
        for i in range(3, 11, 2):
            if len(row) > i:
                item = str_val(row[i])
                if item:
                    extra[current_section].append(item)
data['extra'] = extra

# === BIG GOALS ===
ws = wb['BIG GOALS']
big_goals = [str_val(row[0]) for row in ws.iter_rows(min_row=1, values_only=True) if str_val(row[0])]
data['bigGoals'] = big_goals

# === VERSION LOG ===
ws = wb['VERSION LOG']
version_log = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0] and row[1]:
        version_log.append({
            'date': str(row[0]),
            'version': str_val(row[1]) or '',
            'description': str_val(row[2]) or '',
            'summary': str_val(row[4]) or '',
            'details': str_val(row[5]) or ''
        })
data['versionLog'] = version_log

# === MODULAR GEAR ===
ws = wb['MODULAR GEAR (LEGACY)']
modular = {}
current_section = None
current_prism = None
current_scaffold = None
for row in ws.iter_rows(min_row=8, values_only=True):
    col0 = str_val(row[0])
    col2 = str_val(row[2]) if len(row) > 2 else None
    col3 = str_val(row[3]) if len(row) > 3 else None
    if col0 and col0.isupper() and not any(isinstance(row[i], bool) for i in range(4, 8) if len(row) > i):
        current_section = col0
        current_prism = None
        current_scaffold = None
        if current_section not in modular:
            modular[current_section] = []
        continue
    if current_section is None:
        continue
    if col0:
        current_prism = col0
    if col2:
        current_scaffold = col2
    if col3 and current_prism:
        label = current_prism
        if current_scaffold:
            label += ' + ' + current_scaffold
        label += ' + ' + col3
        modular[current_section].append(label)
data['modularGear'] = modular

# === SETTINGS ===
data['settingsDefinitions'] = {
    'founder': 'Unlock tracking for founder only gear',
    'conclave': 'Include conclave collections',
    'gear': {
        'reactor': 'Reactor/Catalyst on everything',
        'exilus': 'Exilus mod on all gear',
        'shards': '5 Shards in each frame',
        'tauForged': 'Tau-Forged shards',
        'arcaneAdapter': 'Arcane adapter on all gear',
        'maxBuild': 'Min-maxed build on all gear',
        'auraForma': 'Aura Forma on all frames',
        'stanceForma': 'Stance Forma on all melee',
        'ampArcaneAdapter': 'Arcane adapter on all amps',
        'lens': 'Max lens on all gear'
    },
    'incarnon': {'completionist': 'Track incarnon adapters on all weapon variants'},
    'arcane': {'psycho': 'Collect 1 of every arcane rank'},
    'mod': {'hoarder': 'Track mod duplicates at every rank'},
    'railjack': {'partHoarder': 'Track all RJ parts with best stats'},
    'relic': {'hoarder': 'Track upgraded relic variants'},
    'blueprint': {'hoarder': 'Include old/impossible blueprints'},
    'cosmetics': {
        'prime': 'Prime Access accessories',
        'consoleExclusive': 'Console exclusive cosmetics',
        'tennogen': 'Tennogen cosmetics',
        'steamItems': 'Steam exclusive skins',
        'nightwave': 'Nightwave cosmetics',
        'old': 'Old impossible cosmetics',
        'extra': 'Tennocon & event exclusive cosmetics',
        'founder': 'Founder exclusive cosmetics'
    },
    'collectable': {
        'eventLocked': 'Event limited collectables',
        'old': 'Old impossible collectables',
        'prime': 'Primed collectables',
        'consoleExclusive': 'Console exclusive collectables',
        'extra': 'Heavily restricted collectables'
    },
    'decorations': {
        'primeAccess': 'Prime Access decorations',
        'events': 'Event decorations',
        'nightwave': 'Nightwave decorations',
        'old': 'Old impossible decorations',
        'extra': 'Tennocon decorations',
        'founder': 'Founder decorations'
    },
    'codex': {'old': 'Include old impossible codex scans'},
    'market': {'extra': 'Tennocon & event market items'},
    'extra': {'prime': 'Prime items', 'plat': 'Platinum items', 'founder': 'Founder items'}
}

output_path = r'D:\1 Arbeit\4 - Programmieren\2026\wf-tracker\wf-tracker\src\assets\tracker-data.json'
os.makedirs(os.path.dirname(output_path), exist_ok=True)
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print('Done! Summary:')
for k, v in data.items():
    if isinstance(v, list):
        print(f'  {k}: {len(v)} items')
    elif isinstance(v, dict):
        total = sum(len(vv) if isinstance(vv, list) else (sum(len(x) for x in vv.values()) if isinstance(vv, dict) else 1) for vv in v.values())
        print(f'  {k}: {len(v)} sections, ~{total} items')
